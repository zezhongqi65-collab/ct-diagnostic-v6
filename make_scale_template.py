"""
量表数据收集模板生成器。

生成一个带数据校验的 Excel 模板，供前测阶段录入学生数据：
列 = 学生ID + 班级 + A1..A15（每题 1-5）+ 已有成绩。

- A1..A15 列按维度分组配色，表头可直观看出每题属于哪个计算思维维度。
- 已内置数据校验：A1..A15 限 1-5 整数，已有成绩限 0-满分。
- 附「填写说明」sheet，解释每列含义与维度对应关系。

用法:
    python make_scale_template.py [--score-max 100]

输出:
    data/量表数据收集模板.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Windows 控制台默认 GBK，强制 stdout 用 UTF-8，避免中文/符号乱码（Python 3.7+）
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

_PARENT = Path(__file__).resolve().parent
if str(_PARENT) not in sys.path:
    sys.path.insert(0, str(_PARENT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from utils.scale_data import ALL_ITEMS, DIMENSION_ITEMS, SCORE_MAX

# ── 维度 → 表头配色（浅色，便于识别分组）──────────────
DIM_COLORS: dict[str, str] = {
    '分解': 'FFF2CC',
    '抽象': 'D9E1F2',
    '建模': 'E2EFDA',
    '算法设计': 'FCE4D6',
    '评估': 'E4DFEC',
}

MAX_ROWS = 400  # 数据区预留的最大行数（含表头），300+ 学生足够


def _item_to_dim(item: str) -> str:
    """返回题项所属维度名。"""
    for dim, items in DIMENSION_ITEMS.items():
        if item in items:
            return dim
    raise ValueError(f"题项 {item} 未在 DIMENSION_ITEMS 中登记")


def build_template(score_max: float) -> Workbook:
    """构建模板工作簿。"""
    wb = Workbook()

    # ── Sheet 1: 数据录入 ──────────────────────────────
    ws = wb.active
    ws.title = '数据录入'

    headers = ['学生ID', '班级'] + ALL_ITEMS + ['已有成绩']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if header in DIMENSION_ITEMS:
            fill = PatternFill('solid', fgColor=DIM_COLORS[header])
        elif header in ALL_ITEMS:
            dim = _item_to_dim(header)
            fill = PatternFill('solid', fgColor=DIM_COLORS[dim])
        else:
            fill = PatternFill('solid', fgColor='D0CECE')
        cell.fill = fill

    ws.freeze_panes = 'C2'  # 冻结表头与前两列

    # 数据校验：A1..A15 允许填 A-E 字母或 1-5 数字
    first_item_col = 3  # A1 位于第 3 列（第 1 列学生ID、第 2 列班级）
    last_item_col = first_item_col + len(ALL_ITEMS) - 1
    dv_items = DataValidation(
        type='list', formula1='"A,B,C,D,E,1,2,3,4,5"',
        allow_blank=True,
    )
    dv_items.errorTitle = '输入错误'
    dv_items.error = '本题只能填 A-E 或 1-5'
    dv_items.promptTitle = '计算思维量表'
    dv_items.prompt = '填 A-E（A=完全不符合，E=完全符合）或 1-5 分'
    ws.add_data_validation(dv_items)
    dv_items.add(
        f'{get_column_letter(first_item_col)}2:'
        f'{get_column_letter(last_item_col)}{MAX_ROWS}'
    )

    # 数据校验：已有成绩 限 0-满分
    score_col = last_item_col + 1
    dv_score = DataValidation(
        type='decimal', operator='between', formula1='0', formula2=str(score_max),
        allow_blank=True,
    )
    dv_score.errorTitle = '输入错误'
    dv_score.error = f'已有成绩需在 0-{int(score_max)} 之间'
    dv_score.promptTitle = '已有成绩'
    dv_score.prompt = f'请填写学生的已有成绩（满分 {int(score_max)}）'
    ws.add_data_validation(dv_score)
    dv_score.add(f'{get_column_letter(score_col)}2:{get_column_letter(score_col)}{MAX_ROWS}')

    # 列宽
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    for col_idx in range(first_item_col, score_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 6

    # ── Sheet 2: 填写说明 ──────────────────────────────
    guide = wb.create_sheet('填写说明')
    rows = [
        ['计算思维量表前测数据收集模板 — 填写说明'],
        [''],
        ['1. 每行代表一名学生，请勿合并或跳过表头。'],
        ['2. 学生ID：学号或任意唯一编号（例如 20240101）。'],
        ['3. 班级：填写班级，用于后续分组与统计分析（例如 高二1班）。'],
        ['4. A1-A15：计算思维量表题项，可填字母 A-E 或数字 1-5。'],
        ['   A=1分(完全不符合) B=2分 C=3分(一般) D=4分 E=5分(完全符合)。'],
        ['   也可直接填 1-5 的数字，系统会自动识别。'],
        ['5. 已有成绩：学生的已有课程/编程成绩，用于训练诊断模型。'],
        [f'   满分按 {int(score_max)} 分计，系统会自动归一化到 0-100。'],
        [''],
        ['维度与题项对应关系（与诊断系统严格一致）：'],
        ['维度', '题项', '题数'],
    ]
    for dim, items in DIMENSION_ITEMS.items():
        rows.append([dim, '、'.join(items), len(items)])
    rows += [
        [''],
        ['提示：表头颜色相同的题项属于同一计算思维维度。'],
        ['填完后另存为 .xlsx 或 .csv，即可交给预处理脚本生成诊断输入。'],
    ]
    for row in rows:
        guide.append(row)

    guide.column_dimensions['A'].width = 60
    guide.column_dimensions['B'].width = 40
    guide.column_dimensions['C'].width = 8
    guide['A1'].font = Font(bold=True, size=14)
    # 维度对应表标题行加粗
    for cell in guide[9]:  # 第 10 行为「维度/题项/题数」表头
        cell.font = Font(bold=True)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description='生成计算思维量表数据收集模板')
    parser.add_argument('--score-max', type=float, default=SCORE_MAX,
                        help=f'已有成绩满分（默认 {SCORE_MAX}）')
    parser.add_argument('--out', type=str, default=None,
                        help='输出路径（默认 data/量表数据收集模板.xlsx）')
    args = parser.parse_args()

    wb = build_template(args.score_max)

    out_path = Path(args.out) if args.out else (
        Path(__file__).resolve().parent / 'data' / '量表数据收集模板.xlsx'
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f'[OK] 模板已生成: {out_path}')
    print('   维度 -> 题项映射:')
    for dim, items in DIMENSION_ITEMS.items():
        print(f'     {dim}: {", ".join(items)}')


if __name__ == '__main__':
    main()
